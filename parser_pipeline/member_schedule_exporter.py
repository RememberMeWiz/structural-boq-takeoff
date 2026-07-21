"""
Member-Level Schedule Exporter
Parses member size extraction JSON and exports a structured Excel schedule
listing each structural member with its tag, type, occurrence count, dimensions,
length, and computed concrete volume.
"""

import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_member_schedule(json_path: str, output_path: str):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    schedule = data.get("schedule", [])
    members = data.get("members", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Member Takeoff Schedule"
    ws.views.sheetView[0].showGridLines = True

    # Styling
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    double_bottom_border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000")
    )

    # Title Block
    ws["A1"] = "STRUCTURAL MEMBER-LEVEL TAKEOFF SCHEDULE"
    ws["A1"].font = title_font
    ws["A2"] = f"Source CAD Drawing: {os.path.basename(data.get('source_dxf', ''))}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="595959")

    headers = [
        "Item No.", "Member Type", "Mark / Tag", "Occurrences",
        "Width (mm)", "Depth (mm)", "Avg Length (mm)", "Total Length (m)",
        "Est. Volume (cu.m.)", "Extraction Status"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Group members by tag
    tag_members_map = {}
    for m in members:
        tag_members_map.setdefault(m["tag"], []).append(m)

    row_idx = 5
    for item_idx, s in enumerate(schedule, 1):
        tag = s.get("tag")
        m_type = s.get("member_type", "beam").capitalize()
        occ_count = s.get("occurrence_count", 0)

        # Retrieve dimensions and lengths from associated members
        assoc_members = tag_members_map.get(tag, [])
        widths = [m["width_mm"] for m in assoc_members if m.get("width_mm")]
        depths = [m["depth_mm"] for m in assoc_members if m.get("depth_mm")]
        lengths = [m["geometry_length"] for m in assoc_members if m.get("geometry_length")]

        avg_width = sum(widths) / len(widths) if widths else (300.0 if m_type.lower() == "beam" else 350.0)
        avg_depth = sum(depths) / len(depths) if depths else (500.0 if m_type.lower() == "beam" else 350.0)
        avg_len = sum(lengths) / len(lengths) if lengths else 3000.0
        tot_len_m = (avg_len * occ_count) / 1000.0
        est_vol = (tot_len_m * (avg_width / 1000.0) * (avg_depth / 1000.0))

        ws.cell(row=row_idx, column=1, value=item_idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=m_type).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=tag).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=occ_count).number_format = "#,##0"
        
        c_w = ws.cell(row=row_idx, column=5, value=avg_width)
        c_w.number_format = "#,##0"
        c_d = ws.cell(row=row_idx, column=6, value=avg_depth)
        c_d.number_format = "#,##0"
        c_l = ws.cell(row=row_idx, column=7, value=avg_len)
        c_l.number_format = "#,##0.0"
        
        c_tot = ws.cell(row=row_idx, column=8, value=f"=(D{row_idx}*G{row_idx})/1000")
        c_tot.number_format = "#,##0.00"
        
        c_vol = ws.cell(row=row_idx, column=9, value=f"=H{row_idx}*(E{row_idx}/1000)*(F{row_idx}/1000)")
        c_vol.number_format = "#,##0.000"
        
        status_text = "Extracted" if widths and depths else "Geometry Parsed"
        ws.cell(row=row_idx, column=10, value=status_text).alignment = Alignment(horizontal="center")

        for c in range(1, 11):
            ws.cell(row=row_idx, column=c).border = thin_border
            ws.cell(row=row_idx, column=c).font = regular_font

        row_idx += 1

    # Total Row
    ws.cell(row=row_idx, column=3, value="TOTAL").font = bold_font
    ws.cell(row=row_idx, column=4, value=f"=SUM(D5:D{row_idx-1})").font = bold_font
    ws.cell(row=row_idx, column=8, value=f"=SUM(H5:H{row_idx-1})").font = bold_font
    tot_vol_cell = ws.cell(row=row_idx, column=9, value=f"=SUM(I5:I{row_idx-1})")
    tot_vol_cell.font = bold_font
    tot_vol_cell.number_format = "#,##0.000"
    tot_vol_cell.border = double_bottom_border

    # Auto Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Successfully exported Member Schedule: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    json_p = sys.argv[1] if len(sys.argv) > 1 else "outputs/qa/task3_member_sizes_residential.json"
    out_p = sys.argv[2] if len(sys.argv) > 2 else "outputs/takeoff_member_schedule_residential.xlsx"
    export_member_schedule(json_p, out_p)
