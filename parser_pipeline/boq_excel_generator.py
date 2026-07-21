"""
Excel Export Engine for Quantity Takeoff and Bill of Quantities (BOQ).
Generates a multi-sheet Excel workbook (.xlsx) with live inter-sheet formulas
matching the specifications in Section 2.8 and 2.9 of tech_spec.md.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
try:
    from parser_pipeline.fajardo_takeoff_engine import BackupComputationRow, BOQChecklistItem, FajardoTakeoffEngine
except ModuleNotFoundError:
    from fajardo_takeoff_engine import BackupComputationRow, BOQChecklistItem, FajardoTakeoffEngine


def generate_boq_workbook(engine: FajardoTakeoffEngine, output_path: str = "outputs/takeoff_boq_schedule.xlsx"):
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------------
    # STYLES DEFINITION
    # -------------------------------------------------------------------
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="595959")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    double_bottom_border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000")
    )
    
    # -------------------------------------------------------------------
    # SHEET 1: UNIT COST DERIVATION (BASE PRICES)
    # -------------------------------------------------------------------
    ws_rates = wb.active
    ws_rates.title = "Unit Cost Derivation"
    ws_rates.views.sheetView[0].showGridLines = True
    
    ws_rates["A1"] = "UNIT COST DERIVATION & BASE MATERIAL/LABOR RATES"
    ws_rates["A1"].font = title_font
    ws_rates["A2"] = "Editable base parameters referencing regional Philippine construction material & labor costs"
    ws_rates["A2"].font = subtitle_font
    
    rate_headers = ["Resource Code", "Resource / Material Name", "Unit", "Base Unit Cost (PHP)", "Notes / Source"]
    for col_num, h in enumerate(rate_headers, 1):
        cell = ws_rates.cell(row=4, column=col_num, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    rates_data = [
        ("MAT-01", "Cement (40kg bag)", "bag", engine.base_prices.get("Cement (40kg bag)", 310.0), "Fajardo Class A/B Mix Standard"),
        ("MAT-02", "Sand (cu.m.)", "cu.m.", engine.base_prices.get("Sand (cu.m.)", 1200.0), "Washed River Sand"),
        ("MAT-03", "Gravel (cu.m.)", "cu.m.", engine.base_prices.get("Gravel (cu.m.)", 1400.0), "Gravel 3/4\" Crushed"),
        ("MAT-04", "Rebar (per kg)", "kg", engine.base_prices.get("Rebar (per kg)", 45.0), "Deformed Bar Grade 33/40"),
        ("MAT-05", "Tie Wire #16 G.I. (per kg)", "kg", engine.base_prices.get("Tie Wire #16 G.I. (per kg)", 90.0), "15kg per Metric Ton Steel"),
        ("MAT-06", "100mm CHB (per pc)", "pc", engine.base_prices.get("100mm CHB (per pc)", 14.0), "Standard 4\" Block"),
        ("MAT-07", "150mm CHB (per pc)", "pc", engine.base_prices.get("150mm CHB (per pc)", 18.0), "Standard 6\" Block"),
        ("LAB-01", "Concrete Works Labor (per cu.m.)", "cu.m.", engine.base_prices.get("Concrete Works Labor (per cu.m.)", 850.0), "Mixing, Pouring & Compaction"),
        ("LAB-02", "Rebar Works Labor (per kg)", "kg", engine.base_prices.get("Rebar Works Labor (per kg)", 12.0), "Cutting, Bending & Tying"),
        ("LAB-03", "Masonry Works Labor (per sq.m.)", "sq.m.", engine.base_prices.get("Masonry Works Labor (per sq.m.)", 220.0), "Laying & 16mm Plastering"),
    ]
    
    for row_idx, r in enumerate(rates_data, 5):
        ws_rates.cell(row=row_idx, column=1, value=r[0]).alignment = Alignment(horizontal="center")
        ws_rates.cell(row=row_idx, column=2, value=r[1])
        ws_rates.cell(row=row_idx, column=3, value=r[2]).alignment = Alignment(horizontal="center")
        c_val = ws_rates.cell(row=row_idx, column=4, value=r[3])
        c_val.number_format = "#,##0.00"
        c_val.alignment = Alignment(horizontal="right")
        ws_rates.cell(row=row_idx, column=5, value=r[4])
        for c in range(1, 6):
            ws_rates.cell(row=row_idx, column=c).border = thin_border
            ws_rates.cell(row=row_idx, column=c).font = regular_font

    # -------------------------------------------------------------------
    # SHEET 2: BACK-UP COMPUTATION (ELEMENT TAKEOFF)
    # -------------------------------------------------------------------
    ws_backup = wb.create_sheet(title="Back-Up Computation")
    ws_backup.views.sheetView[0].showGridLines = True
    
    ws_backup["A1"] = "DETAILED BACK-UP COMPUTATION SHEET"
    ws_backup["A1"].font = title_font
    ws_backup["A2"] = "Element-level geometric takeoff & quantity derivations traceable to drawing sheets"
    ws_backup["A2"].font = subtitle_font
    
    backup_headers = [
        "Work Section", "Item Code", "Element & Description", "Location Description",
        "Drawing Ref", "Length / Area (m)", "Width (m)", "Height / Thick (m)", "Count",
        "Takeoff Qty", "Unit", "Unit Cost (PHP)", "Amount (PHP)", "Status"
    ]
    for col_num, h in enumerate(backup_headers, 1):
        cell = ws_backup.cell(row=4, column=col_num, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    curr_row = 5
    for r in engine.backup_rows:
        ws_backup.cell(row=curr_row, column=1, value=r.work_section)
        ws_backup.cell(row=curr_row, column=2, value=r.item_code).alignment = Alignment(horizontal="center")
        ws_backup.cell(row=curr_row, column=3, value=r.description)
        ws_backup.cell(row=curr_row, column=4, value=r.location_description)
        ws_backup.cell(row=curr_row, column=5, value=r.drawing_ref).alignment = Alignment(horizontal="center")
        
        ws_backup.cell(row=curr_row, column=6, value=r.length_or_area).number_format = "#,##0.00"
        ws_backup.cell(row=curr_row, column=7, value=r.width).number_format = "#,##0.00"
        ws_backup.cell(row=curr_row, column=8, value=r.height_or_thickness).number_format = "#,##0.00"
        ws_backup.cell(row=curr_row, column=9, value=r.count).number_format = "#,##0"
        
        # Takeoff Qty formula or value
        q_cell = ws_backup.cell(row=curr_row, column=10, value=r.quantity)
        q_cell.number_format = "#,##0.00"
        q_cell.alignment = Alignment(horizontal="right")
        
        ws_backup.cell(row=curr_row, column=11, value=r.unit).alignment = Alignment(horizontal="center")
        
        # Keep pricing aligned with the engine's concrete and mortar classes.
        unit_cost = engine.unit_cost_for_row(r)
            
        uc_cell = ws_backup.cell(row=curr_row, column=12, value=unit_cost)
        uc_cell.number_format = "#,##0.00"
        
        # Amount formula: Takeoff Qty * Unit Cost
        amt_cell = ws_backup.cell(row=curr_row, column=13, value=f"=J{curr_row}*L{curr_row}")
        amt_cell.number_format = "#,##0.00"
        amt_cell.alignment = Alignment(horizontal="right")
        
        ws_backup.cell(row=curr_row, column=14, value=r.status).alignment = Alignment(horizontal="center")
        
        for c in range(1, 15):
            ws_backup.cell(row=curr_row, column=c).border = thin_border
            ws_backup.cell(row=curr_row, column=c).font = regular_font
            
        curr_row += 1

    # Total row for Back-Up Computation
    ws_backup.cell(row=curr_row, column=3, value="TOTAL BACK-UP COMPUTATION AMOUNT (PHP)").font = bold_font
    total_amt_cell = ws_backup.cell(row=curr_row, column=13, value=f"=SUM(M5:M{curr_row-1})")
    total_amt_cell.font = bold_font
    total_amt_cell.number_format = "#,##0.00"
    total_amt_cell.border = double_bottom_border

    # -------------------------------------------------------------------
    # SHEET 3: CHECKLIST / BOQ SUMMARY (ROLLED UP)
    # -------------------------------------------------------------------
    ws_boq = wb.create_sheet(title="Checklist BOQ Summary")
    ws_boq.views.sheetView[0].showGridLines = True
    
    ws_boq["A1"] = "BILL OF QUANTITIES (BOQ) CHECKLIST SUMMARY"
    ws_boq["A1"].font = title_font
    ws_boq["A2"] = "Itemized quantity rollup and cost summary for Concrete, Rebar & Masonry Trades"
    ws_boq["A2"].font = subtitle_font
    
    boq_headers = ["Item No.", "Item Code", "Description", "Unit", "Takeoff Qty", "Blended Unit Cost (PHP)", "Total Amount (PHP)", "Status"]
    for col_num, h in enumerate(boq_headers, 1):
        cell = ws_boq.cell(row=4, column=col_num, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    boq_items = engine.summarize_to_boq()
    boq_row = 5
    for b in boq_items:
        ws_boq.cell(row=boq_row, column=1, value=b.item_no).alignment = Alignment(horizontal="center")
        ws_boq.cell(row=boq_row, column=2, value=b.item_code).alignment = Alignment(horizontal="center")
        ws_boq.cell(row=boq_row, column=3, value=b.description)
        ws_boq.cell(row=boq_row, column=4, value=b.unit).alignment = Alignment(horizontal="center")
        
        # Live Excel formula linking back to Back-Up Computation Qty sum
        qty_formula = f"=SUMIF('Back-Up Computation'!B:B, B{boq_row}, 'Back-Up Computation'!J:J)"
        q_cell = ws_boq.cell(row=boq_row, column=5, value=qty_formula)
        q_cell.number_format = "#,##0.00"
        q_cell.alignment = Alignment(horizontal="right")
        
        # Blended Unit Cost formula: Total Amount / Total Qty
        uc_formula = f"=IF(E{boq_row}>0, G{boq_row}/E{boq_row}, 0)"
        uc_cell = ws_boq.cell(row=boq_row, column=6, value=uc_formula)
        uc_cell.number_format = "#,##0.00"
        uc_cell.alignment = Alignment(horizontal="right")
        
        # Live Excel formula linking to Back-Up Computation Amount sum
        amt_formula = f"=SUMIF('Back-Up Computation'!B:B, B{boq_row}, 'Back-Up Computation'!M:M)"
        amt_cell = ws_boq.cell(row=boq_row, column=7, value=amt_formula)
        amt_cell.number_format = "#,##0.00"
        amt_cell.alignment = Alignment(horizontal="right")
        
        ws_boq.cell(row=boq_row, column=8, value=b.status).alignment = Alignment(horizontal="center")
        
        for c in range(1, 9):
            ws_boq.cell(row=boq_row, column=c).border = thin_border
            ws_boq.cell(row=boq_row, column=c).font = regular_font
            
        boq_row += 1

    # Total BOQ Summary row
    ws_boq.cell(row=boq_row, column=3, value="GRAND TOTAL PROJECT DIRECT COST (PHP)").font = bold_font
    total_boq_cell = ws_boq.cell(row=boq_row, column=7, value=f"=SUM(G5:G{boq_row-1})")
    total_boq_cell.font = bold_font
    total_boq_cell.number_format = "#,##0.00"
    total_boq_cell.border = double_bottom_border

    # Adjust column widths automatically for all sheets
    for sheet in [ws_rates, ws_backup, ws_boq]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Successfully generated BOQ Workbook: {output_path}")
    return output_path


if __name__ == "__main__":
    try:
        from parser_pipeline.fajardo_takeoff_engine import TakeoffElement
    except ModuleNotFoundError:
        from fajardo_takeoff_engine import TakeoffElement
    
    eng = FajardoTakeoffEngine()
    
    # Load sample elements
    test_elements = [
        TakeoffElement("e1", "footing", "F-1", "Grid 1-A to 4-D", "S-1", 1.5, 1.5, 0.4, 12, "Class A", [{"diameter": 16, "count": 10, "length": 1.7}]),
        TakeoffElement("e2", "column", "C-1", "Ground to 2nd Floor", "S-1", 0.35, 0.35, 3.2, 12, "Class A", [{"diameter": 20, "count": 8, "length": 3.8}, {"diameter": 10, "count": 22, "length": 1.3}]),
        TakeoffElement("e3", "beam", "2B-1", "2nd Floor Framing", "S-2", 6.0, 0.30, 0.50, 8, "Class A", [{"diameter": 20, "count": 6, "length": 6.8}, {"diameter": 10, "count": 35, "length": 1.5}]),
        TakeoffElement("e4", "chb_wall", "W-1", "Exterior Perimeter Wall", "A-1", 35.0, 0.15, 3.0, 1, chb_thickness="150mm", plaster_faces=2),
    ]
    for elem in test_elements:
        if elem.element_type in ["footing", "column", "beam", "slab"]:
            eng.process_concrete_element(elem)
            if elem.rebar_specs:
                eng.process_rebar_specs(elem)
        elif elem.element_type == "chb_wall":
            eng.process_masonry_element(elem)
            
    generate_boq_workbook(eng, "outputs/takeoff_boq_schedule.xlsx")
