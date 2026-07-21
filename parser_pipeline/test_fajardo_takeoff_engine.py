"""
Unit tests for Fajardo Takeoff Engine & BOQ Excel Generator.

Phase 3 QA Verification pass: this suite checks both internal consistency
(the engine agrees with itself) AND compliance with tech_spec.md v2.0
(the engine agrees with the approved Fajardo Formula Library figures).
"""

import unittest
import os
import openpyxl
from parser_pipeline.fajardo_takeoff_engine import (
    FajardoTakeoffEngine,
    TakeoffElement,
    CONCRETE_MIX_FACTORS,
    REBAR_UNIT_WEIGHTS,
    CHB_COUNT_PER_SQM,
    MORTAR_CLASS_CEMENT_BAGS_PER_M3,
    MORTAR_CLASS_B_FACTORS,
    PLASTER_CLASS_B_FACTORS_PER_FACE,
    QA_DIVERGENCE_THRESHOLD,
    mortar_factors,
    plaster_factors,
)
from parser_pipeline.boq_excel_generator import generate_boq_workbook


class TestFajardoTakeoffEngine(unittest.TestCase):

    def setUp(self):
        self.engine = FajardoTakeoffEngine()

    # ------------------------------------------------------------------
    # Concrete Works
    # ------------------------------------------------------------------

    def test_concrete_footing_volume(self):
        elem = TakeoffElement(
            element_id="f1", element_type="footing", label="F-1",
            location="Grid A-1", drawing_ref="S-1",
            length=1.5, width=1.5, height_or_thickness=0.4, count=10,
            concrete_class="Class A"
        )
        vol = self.engine.process_concrete_element(elem)
        expected_vol = 1.5 * 1.5 * 0.4 * 10  # 9.0 cu.m.
        self.assertAlmostEqual(vol, expected_vol, places=3)
        self.assertEqual(len(self.engine.backup_rows), 1)
        self.assertEqual(self.engine.backup_rows[0].item_code, "CON-1.1")
        self.assertEqual(self.engine.backup_rows[0].quantity, 9.0)
        self.assertEqual(len(self.engine.calculation_checks), 1)
        self.assertFalse(self.engine.calculation_checks[0].warning)

    def test_concrete_volume_and_linear_meter_methods_agree_by_construction(self):
        # For rectangular members, V=L*W*H*N and section-area*linear-length
        # are algebraically identical, so the dual check should never
        # legitimately flag a straight rectangular member -- confirms the
        # cross-check is wired correctly rather than silently disabled.
        elem = TakeoffElement(
            element_id="b1", element_type="beam", label="B-1",
            location="Test", drawing_ref="S-1",
            length=6.0, width=0.3, height_or_thickness=0.5, count=8,
        )
        self.engine.process_concrete_element(elem)
        check = self.engine.calculation_checks[0]
        self.assertEqual(check.divergence_ratio, 0.0)
        self.assertFalse(check.warning)

    # ------------------------------------------------------------------
    # Steel Reinforcement
    # ------------------------------------------------------------------

    def test_rebar_weight_calculation(self):
        elem = TakeoffElement(
            element_id="c1", element_type="column", label="C-1",
            location="Grid B-2", drawing_ref="S-1",
            length=3.0, width=0.3, height_or_thickness=0.3, count=1,
            rebar_specs=[{"diameter": 16, "count": 4, "length": 3.5}]
        )
        wt = self.engine.process_rebar_specs(elem)
        # 4 bars * 3.5m = 14.0m; 14.0m * 1.578 kg/m = 22.092 kg
        expected_wt = 4 * 3.5 * REBAR_UNIT_WEIGHTS[16]
        self.assertAlmostEqual(wt, expected_wt, places=2)
        self.assertEqual(len(self.engine.backup_rows), 2)  # rebar row + tie wire row
        self.assertEqual(self.engine.backup_rows[0].unit, "kg")
        self.assertEqual(len(self.engine.calculation_checks), 1)
        self.assertLess(self.engine.calculation_checks[0].divergence_ratio, QA_DIVERGENCE_THRESHOLD)

    def test_rebar_unit_weights_match_pns49_astm_a615_table(self):
        # tech_spec.md v2.0 Table 2.6.2
        expected = {10: 0.617, 12: 0.888, 16: 1.578, 20: 2.466, 25: 3.853, 28: 4.834, 32: 6.313}
        self.assertEqual(REBAR_UNIT_WEIGHTS, expected)

    def test_gi_tie_wire_factor_applied_to_columns_and_beams(self):
        elem = TakeoffElement(
            element_id="c2", element_type="column", label="C-2",
            location="Test", drawing_ref="S-1",
            length=3.0, width=0.3, height_or_thickness=0.3, count=1,
            rebar_specs=[{"diameter": 20, "count": 8, "length": 3.8}]
        )
        total_rebar_kg = self.engine.process_rebar_specs(elem)
        tie_wire_row = self.engine.backup_rows[-1]
        self.assertEqual(tie_wire_row.item_code, "REB-2.0")
        self.assertAlmostEqual(tie_wire_row.quantity, round(total_rebar_kg * 0.015, 2), places=2)

    # ------------------------------------------------------------------
    # Masonry (CHB) Works
    # ------------------------------------------------------------------

    def test_masonry_chb_wall_area(self):
        elem = TakeoffElement(
            element_id="w1", element_type="chb_wall", label="W-1",
            location="Exterior Wall", drawing_ref="A-1",
            length=10.0, width=0.15, height_or_thickness=3.0, count=1,
            chb_thickness="150mm", plaster_faces=2
        )
        area = self.engine.process_masonry_element(elem)
        self.assertAlmostEqual(area, 30.0, places=2)
        self.assertEqual(len(self.engine.backup_rows), 2)  # CHB wall row + plaster row
        self.assertEqual(self.engine.backup_rows[0].item_code, "MAS-3.2")
        self.assertEqual(self.engine.backup_rows[1].quantity, 60.0)  # 30 sqm * 2 faces
        self.assertFalse(self.engine.calculation_checks[0].warning)

    def test_masonry_openings_are_deducted_from_chb_and_plaster(self):
        elem = TakeoffElement(
            element_id="w-openings", element_type="chb_wall", label="W-2",
            location="Exterior Wall", drawing_ref="A-1", length=10.0, width=0.15,
            height_or_thickness=3.0, count=2, chb_thickness="100mm",
            opening_area=1.5, plaster_faces=2,
        )
        area = self.engine.process_masonry_element(elem)
        self.assertEqual(area, 57.0)  # (10 x 3 x 2) - (1.5 x 2)
        self.assertEqual(self.engine.backup_rows[0].quantity, 57.0)
        self.assertEqual(self.engine.backup_rows[1].quantity, 114.0)

    def test_negative_opening_area_raises(self):
        elem = TakeoffElement(
            element_id="w-neg", element_type="chb_wall", label="W-N",
            location="Test", drawing_ref="A-1", length=5.0, width=0.15,
            height_or_thickness=3.0, count=1, chb_thickness="100mm",
            opening_area=-1.0,
        )
        with self.assertRaises(ValueError):
            self.engine.process_masonry_element(elem)

    def test_opening_area_exceeding_gross_area_raises(self):
        elem = TakeoffElement(
            element_id="w-huge-opening", element_type="chb_wall", label="W-H",
            location="Test", drawing_ref="A-1", length=2.0, width=0.15,
            height_or_thickness=2.0, count=1, chb_thickness="100mm",
            opening_area=10.0,
        )
        with self.assertRaises(ValueError):
            self.engine.process_masonry_element(elem)

    def test_chb_count_divergence_is_flagged(self):
        elem = TakeoffElement(
            element_id="w-qa", element_type="chb_wall", label="W-3",
            location="Test", drawing_ref="A-1", length=1.0, width=0.10,
            height_or_thickness=1.0, count=1, chb_thickness="100mm",
        )
        self.engine.process_masonry_element(elem)
        check = self.engine.calculation_checks[0]
        self.assertTrue(check.warning)
        self.assertGreater(check.divergence_ratio, QA_DIVERGENCE_THRESHOLD)
        self.assertIn("QA warning", self.engine.backup_rows[0].status)

    def test_plastering_has_its_own_dual_cross_check(self):
        # tech_spec.md §2.5.3 Dual Cross-Checks item 3: "Plastering: Volume
        # Method vs. Area Method". Previously the engine only cross-checked
        # CHB count and silently reused that result for the plaster row.
        elem = TakeoffElement(
            element_id="w-plaster-check", element_type="chb_wall", label="W-PC",
            location="Test", drawing_ref="A-1", length=8.0, width=0.15,
            height_or_thickness=2.8, count=1, chb_thickness="150mm", plaster_faces=2,
        )
        self.engine.process_masonry_element(elem)
        check_names = [c.check_name for c in self.engine.calculation_checks]
        self.assertEqual(len(self.engine.calculation_checks), 2)
        self.assertTrue(any("CHB count" in n for n in check_names))
        self.assertTrue(any("Plastering" in n for n in check_names))

    # ------------------------------------------------------------------
    # Mortar / Plaster Class Scale (distinct from concrete mix classes)
    # ------------------------------------------------------------------

    def test_mortar_classes_are_separate_from_concrete_mix_classes(self):
        class_a = mortar_factors("100mm", "Class A")
        class_b = mortar_factors("100mm", "Class B")
        class_d = mortar_factors("100mm", "Class D")
        self.assertEqual(set(MORTAR_CLASS_CEMENT_BAGS_PER_M3), {"Class A", "Class B", "Class C", "Class D"})
        self.assertGreater(class_a["cement_bags"], class_b["cement_bags"])
        self.assertLess(class_d["cement_bags"], class_b["cement_bags"])
        self.assertGreater(plaster_factors("Class A")["cement_bags"], plaster_factors("Class D")["cement_bags"])

    def test_mortar_class_b_baseline_matches_spec(self):
        # tech_spec.md v2.0 §2.6.3: "Laying Mortar & Cell Fill per 1 sq.m. Wall (Class B Mortar)"
        self.assertEqual(MORTAR_CLASS_B_FACTORS["100mm"], {"cement_bags": 0.522, "sand_m3": 0.0435})
        self.assertEqual(MORTAR_CLASS_B_FACTORS["150mm"], {"cement_bags": 1.010, "sand_m3": 0.0840})

    def test_plaster_class_b_baseline_matches_spec(self):
        # tech_spec.md v2.0 §2.6.3: "Plastering per 1 sq.m. Wall Surface (16mm, Class B Plaster) - One Face"
        self.assertEqual(PLASTER_CLASS_B_FACTORS_PER_FACE["16mm"], {"cement_bags": 0.192, "sand_m3": 0.016})

    def test_mortar_class_scale_matches_published_cement_bag_ratios(self):
        # tech_spec.md v2.0 §2.6.3 table gives exact per-1cu.m. cement bags
        # for each class: A=18.0, B=12.0, C=9.0, D=7.5 (40kg bags).
        # The scaled per-sq.m. factor for each class should carry the same
        # ratio relative to Class B.
        base = mortar_factors("150mm", "Class B")["cement_bags"]
        for cls, spec_bags in [("Class A", 18.0), ("Class C", 9.0), ("Class D", 7.5)]:
            expected_ratio = spec_bags / 12.0
            actual = mortar_factors("150mm", cls)["cement_bags"]
            self.assertAlmostEqual(actual / base, expected_ratio, places=4)

    def test_selected_mortar_class_changes_chb_wall_rate(self):
        class_a = TakeoffElement(
            element_id="w-a", element_type="chb_wall", label="W-A", location="Test",
            drawing_ref="A-1", length=10, width=.15, height_or_thickness=3,
            chb_thickness="150mm", mortar_class="Class A",
        )
        class_d = TakeoffElement(
            element_id="w-d", element_type="chb_wall", label="W-D", location="Test",
            drawing_ref="A-1", length=10, width=.15, height_or_thickness=3,
            chb_thickness="150mm", mortar_class="Class D",
        )
        self.engine.process_masonry_element(class_a)
        class_a_rate = self.engine.unit_cost_for_row(self.engine.backup_rows[0])
        self.engine.process_masonry_element(class_d)
        class_d_rate = self.engine.unit_cost_for_row(self.engine.backup_rows[2])
        self.assertGreater(class_a_rate, class_d_rate)

    def test_selected_plaster_class_changes_plaster_rate(self):
        elem = TakeoffElement(
            element_id="w-plaster", element_type="chb_wall", label="W-P", location="Test",
            drawing_ref="A-1", length=10, width=.15, height_or_thickness=3,
            chb_thickness="150mm", plaster_faces=2, plaster_class="Class A",
        )
        self.engine.process_masonry_element(elem)
        class_a_rate = self.engine.unit_cost_for_row(self.engine.backup_rows[1])
        elem.plaster_class = "Class D"
        self.engine.process_masonry_element(elem)
        class_d_rate = self.engine.unit_cost_for_row(self.engine.backup_rows[3])
        self.assertGreater(class_a_rate, class_d_rate)

    # ------------------------------------------------------------------
    # QA Divergence Threshold Behavior
    # ------------------------------------------------------------------

    def test_divergence_exactly_at_threshold_does_not_warn(self):
        elem = TakeoffElement(element_id="qa1", element_type="footing", label="X", location="T", drawing_ref="S-1",
                               length=1, width=1, height_or_thickness=1, count=1)
        primary, secondary = 1.0, 0.98  # divergence = |1.0-0.98| / max(1.0, 0.98) = exactly 0.02
        check = self.engine._record_check(elem, "boundary test", primary, secondary)
        self.assertAlmostEqual(check.divergence_ratio, QA_DIVERGENCE_THRESHOLD, places=6)
        self.assertFalse(check.warning)  # strictly greater-than, not >=

    def test_divergence_just_above_threshold_warns(self):
        elem = TakeoffElement(element_id="qa2", element_type="footing", label="X", location="T", drawing_ref="S-1",
                               length=1, width=1, height_or_thickness=1, count=1)
        check = self.engine._record_check(elem, "boundary test", 1.0, 0.975)
        self.assertGreater(check.divergence_ratio, QA_DIVERGENCE_THRESHOLD)
        self.assertTrue(check.warning)

    def test_qa_warnings_property_filters_only_warnings(self):
        clean = TakeoffElement(element_id="clean", element_type="footing", label="F-C", location="T",
                                drawing_ref="S-1", length=1.0, width=1.0, height_or_thickness=1.0, count=1)
        divergent = TakeoffElement(element_id="bad", element_type="chb_wall", label="W-Bad", location="T",
                                    drawing_ref="A-1", length=1.0, width=0.1, height_or_thickness=1.0, count=1,
                                    chb_thickness="100mm")
        self.engine.process_concrete_element(clean)
        self.engine.process_masonry_element(divergent)
        self.assertTrue(all(c.warning for c in self.engine.qa_warnings))
        self.assertLess(len(self.engine.qa_warnings), len(self.engine.calculation_checks))

    # ------------------------------------------------------------------
    # Excel Export
    # ------------------------------------------------------------------

    def test_boq_excel_generation(self):
        elem = TakeoffElement(
            element_id="b1", element_type="beam", label="B-1",
            location="2nd Floor", drawing_ref="S-2",
            length=5.0, width=0.3, height_or_thickness=0.4, count=2,
            concrete_class="Class A"
        )
        self.engine.process_concrete_element(elem)
        output_file = "outputs/test_takeoff_boq.xlsx"
        path = generate_boq_workbook(self.engine, output_file)

        self.assertTrue(os.path.exists(path))

        wb = openpyxl.load_workbook(path, data_only=False)
        self.assertIn("Unit Cost Derivation", wb.sheetnames)
        self.assertIn("Back-Up Computation", wb.sheetnames)
        self.assertIn("Checklist BOQ Summary", wb.sheetnames)

        wb.close()
        if os.path.exists(output_file):
            os.remove(output_file)


if __name__ == "__main__":
    unittest.main()
