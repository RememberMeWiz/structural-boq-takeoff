#!/usr/bin/env python3
"""
takeoff_report_generator.py

Task 2 (part 2) — agent: takeoff-report-developer

Objective:
    Consume the layer classification (Task 1) and the geometry aggregation
    (dxf_geometry_engine.py) to produce the actual quantity takeoff
    deliverable: a beam/column schedule with total lengths, entity counts,
    and optional cost-database hooks — written to CSV and/or Excel.

Usage:
    python takeoff_report_generator.py drawing.dxf classification.json \\
        -o takeoff_schedule.xlsx \\
        --cost-table cost_table.json

    # CSV only, no cost data yet:
    python takeoff_report_generator.py drawing.dxf classification.json \\
        -o takeoff_schedule.csv

Cost table format (optional, --cost-table):
    {
      "default": {"beam": 12.50, "column": 18.75},
      "layers":  {"S-BEAM-CONC": 15.00}
    }
    "default" gives a per-length unit cost by label (beam/column).
    "layers" optionally overrides that for specific layer names.
    Units for unit cost must match the drawing's length units (see the
    "units" note printed at runtime) — this script does no unit conversion.

    If no cost table is given, cost fields are left blank/null in the
    output rather than guessed — the schedule is still fully usable for
    lengths/counts, and cost can be filled in later once real pricing
    data is available.
"""

import argparse
import csv
import json
import sys

from dxf_geometry_engine import (
    compute_entity_lengths,
    aggregate_by_layer,
    classification_list_to_map,
    get_drawing_units,
)


def load_classification(path: str) -> dict:
    with open(path) as f:
        data = json.load(f)
    return classification_list_to_map(data)


def load_cost_table(path: str | None) -> dict:
    if not path:
        return {}
    with open(path) as f:
        return json.load(f)


def unit_cost_for(layer: str, label: str, cost_table: dict):
    """Look up a per-length unit cost: layer-specific override first,
    falling back to the label default, or None if no cost data exists."""
    if not cost_table:
        return None
    layer_overrides = cost_table.get("layers", {})
    if layer in layer_overrides:
        return layer_overrides[layer]
    return cost_table.get("default", {}).get(label)


def build_schedule_rows(aggregates, cost_table: dict) -> list[dict]:
    """Turn LayerAggregate objects into flat report rows, one per layer,
    with cost fields populated only where cost data is actually available."""
    rows = []
    for agg in aggregates:
        cost_per_unit = unit_cost_for(agg.layer, agg.label, cost_table)
        total_cost = (
            round(agg.total_length * cost_per_unit, 2)
            if cost_per_unit is not None else None
        )
        rows.append({
            "layer": agg.layer,
            "label": agg.label,
            "entity_count": agg.entity_count,
            "total_length": round(agg.total_length, 3),
            "unit_cost": cost_per_unit,
            "total_cost": total_cost,
        })

    # Add a grand-total row per label (beam / column), which is usually
    # the number someone actually wants at a glance.
    totals_by_label: dict[str, dict] = {}
    for row in rows:
        t = totals_by_label.setdefault(row["label"], {
            "entity_count": 0, "total_length": 0.0, "total_cost": 0.0, "has_cost": False
        })
        t["entity_count"] += row["entity_count"]
        t["total_length"] += row["total_length"]
        if row["total_cost"] is not None:
            t["total_cost"] += row["total_cost"]
            t["has_cost"] = True

    for label, t in totals_by_label.items():
        rows.append({
            "layer": f"TOTAL ({label})",
            "label": label,
            "entity_count": t["entity_count"],
            "total_length": round(t["total_length"], 3),
            "unit_cost": None,
            "total_cost": round(t["total_cost"], 2) if t["has_cost"] else None,
        })

    return rows


def write_csv(rows: list[dict], path: str):
    fieldnames = ["layer", "label", "entity_count", "total_length", "unit_cost", "total_cost"]
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict], path: str, units: str):
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        print(
            "openpyxl not installed — skipping .xlsx output. "
            "Install with: pip install openpyxl --break-system-packages",
            file=sys.stderr,
        )
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Takeoff Schedule"

    headers = ["Layer", "Label", "Count", f"Total Length ({units})", "Unit Cost", "Total Cost"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        is_total = row["layer"].startswith("TOTAL")
        ws.append([
            row["layer"], row["label"], row["entity_count"],
            row["total_length"], row["unit_cost"], row["total_cost"],
        ])
        if is_total:
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 3

    wb.save(path)
    return True


def main():
    parser = argparse.ArgumentParser(description="Generate a beam/column quantity takeoff schedule from a DXF")
    parser.add_argument("dxf_path", help="Path to input DXF file")
    parser.add_argument("classification_path", help="Path to Task 1 classification JSON output")
    parser.add_argument("-o", "--output", default="takeoff_schedule.csv",
                         help="Output path (.csv or .xlsx, default: takeoff_schedule.csv)")
    parser.add_argument("--cost-table", default=None,
                         help="Optional JSON cost table for unit-cost lookups (see module docstring)")
    args = parser.parse_args()

    print(f"Reading DXF: {args.dxf_path}")
    entity_lengths = compute_entity_lengths(args.dxf_path)
    units = get_drawing_units(args.dxf_path)
    print(f"Found {len(entity_lengths)} LINE/LWPOLYLINE entities. Drawing units: {units}")
    if units == "Unitless":
        print("WARNING: drawing has no $INSUNITS set — verify lengths manually against a known dimension.",
              file=sys.stderr)

    classification = load_classification(args.classification_path)
    aggregates = aggregate_by_layer(entity_lengths, classification)
    print(f"Aggregated {len(aggregates)} beam/column layer(s).")

    cost_table = load_cost_table(args.cost_table)
    rows = build_schedule_rows(aggregates, cost_table)

    if args.output.lower().endswith(".xlsx"):
        ok = write_xlsx(rows, args.output, units)
        if not ok:
            fallback = args.output.rsplit(".", 1)[0] + ".csv"
            write_csv(rows, fallback)
            print(f"Wrote CSV fallback instead: {fallback}")
        else:
            print(f"Wrote schedule to {args.output}")
    else:
        write_csv(rows, args.output)
        print(f"Wrote schedule to {args.output}")


if __name__ == "__main__":
    main()
